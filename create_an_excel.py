import openpyxl
from openpyxl.styles import Font

data_frame = openpyxl.Workbook()

data_frame.remove(data_frame.active)

sheet_1 = data_frame.create_sheet("FedEx zip ranges", 0)
sheet_2 = data_frame.create_sheet("UPS zip ranges")

sheet_2.insert_rows(0)
sheet_2["A1"].value = "UPS zone ranges"
sheet_2["B1"].value = "zip from"
sheet_2["C1"].value = "zip to"

sheet_1.insert_rows(0)
sheet_1["A1"].value = "FedEx zone ranges"
sheet_1["B1"].value = "Links to download"
sheet_1["E1"].value = "https://www.fedex.com/ratetools/documents2/"
sheet_1["F1"].value = ".csv"
sheet_1["G1"].value = "=CONCATENATE(E1,A2,F1)"

sheet_1.auto_filter.ref = "A1:C999"
sheet_2.auto_filter.ref = "A1:C999"

font = Font(b=True, size=13, color="050505")

sheet_2["A1"].font = font
sheet_2["B1"].font = font
sheet_2["C1"].font = font

sheet_1["A1"].font = font
sheet_1["B1"].font = font
